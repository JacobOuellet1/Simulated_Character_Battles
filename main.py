# main.py
from copy import deepcopy
from tkinter import *
from tkinter import ttk
import openpyxl

def main():
	# Mock battle
	root = Tk()
	root.title("AI Fight Simulator")

	# Add a grid
	mainframe = Frame(root)
	mainframe.grid(column=0, row=0, sticky=(N, W, E, S))
	mainframe.columnconfigure(0, weight=1)
	mainframe.rowconfigure(0, weight=1)
	mainframe.pack(pady=100, padx=100)

	# Dictionary with options
	character_list = []
	workbook = openpyxl.load_workbook("web_scraper/Character_Database_Standardized.xlsx")
	worksheet = workbook.active
	i = 0
	for row in worksheet.iter_rows():
		if i >= 1:  # skip header of excel sheet
			cell = row[0].value  # get the character name
			cell2 = row[2].value	# get their planet/universe e.g movie version vs character version vs different versions of the same hero
			if cell2 == '-':
				cell2 = 'N/A'
			character_list.append(cell + " | Universe: " + cell2)
		i += 1
	workbook.close()

	popupMenu = ttk.Combobox(mainframe, values=character_list, width=50)
	popupMenu2 = ttk.Combobox(mainframe, values=character_list, width=50)
	Label(mainframe, text="Choose fighter 1").grid(row=1, column=1)
	Label(mainframe, text="Choose fighter 2").grid(row=1, column=3)
	popupMenu.grid(row=2, column=1)
	popupMenu2.grid(row=2, column=3)

	# create button, link it
	def clickRunFight():
		print(popupMenu.get())
		print(popupMenu2.get())

		idx_char_1 = character_list.index(popupMenu.get()) + 2  # add one because of the header row
		idx_char_2 = character_list.index(popupMenu2.get()) + 2  # add one because of the header row and add another one because cell function starts index at 1

		wb = openpyxl.load_workbook("web_scraper/Character_Database_Standardized.xlsx")
		ws = wb.active

		char_1_data = []
		char_2_data = []
		for i in range(1, 15):
			if i != 3:  # skip the planet of the character found at index 2
				if i > 5:  # numbers so convert to float
					char_1_data.append(float(ws.cell(idx_char_1, i).value))
				else:
					char_1_data.append(ws.cell(idx_char_1, i).value)
		for i in range(1, 15):
			if i != 3:  # skip the planet of the character found at index 2
				if i > 5:  # numbers so convert to float
					char_2_data.append(float(ws.cell(idx_char_2, i).value))
				else:
					char_2_data.append(ws.cell(idx_char_2, i).value)

		player_one = Player(*char_1_data)
		player_two = Player(*char_2_data)

		initial_game = Game(player_one, player_two)

		initial_game.p1.print_stats()
		initial_game.p2.print_stats()

		print("A chance encounter occurs, and the following battle takes place:\n")

		the_game = ExpectimaxHeuristic(initial_game)

		the_game.emulate()

	Btn = ttk.Button(mainframe, text="Fight!", command=clickRunFight)

	# place button
	Btn.grid(row=3, column=3)

	root.mainloop()

	"""
	Win condition:
		Health points of enemy depleted
	Draw condition:
		Neither opponent is defeated by turn X (decided arbitrarily by the computer's abilities)
	"""

class Player():
	def __init__(self, super_name, real_name, gender, species, height, weight, inte, stre, spee, dura, powe, comb, p_type):
		self.super_name = super_name
		self.real_name = real_name
		self.gender = gender
		self.species = species
		self.height = height
		self.weight = weight
		self.intelligence = inte
		self.strength = stre
		self.speed = spee
		self.durability = dura
		self.power = powe
		self.combat = comb
		self.player_type = p_type # 0 for minimizer, 1 for maximizer
		self.health = round((self.durability * 0.66 + self.strength * 0.16 + self.power* 0.16) * 15, 1)
		self.blocking = 0
	def print_stats(self):
		print("Player:\t" + str(self.super_name))
		print("Intelligence:\t" + str(self.intelligence))
		print("Strength:\t" + str(self.strength))
		print("Speed:\t\t" + str(self.speed))
		print("Durability:\t" + str(self.durability))
		print("Power:\t\t" + str(self.power))
		print("Combat:\t\t" + str(self.combat))
		print("Player Type:\t" + ("Maximizer" if self.player_type else "Minimizer"))
		print("Blocking:\t" + ("True" if self.blocking else "False"))
		print("Health:\t\t" + str(self.health) + "\n")
	def strike(self, target, blocking):
		self.blocking = blocking
		damage = (self.strength * 0.33 + self.combat * 0.33 + self.power * 0.33) * (3 if self.blocking else 5)
		if target.blocking:
			ratio = (0.8 * (self.speed * 0.66 + self.combat * 0.16 + self.intelligence * 0.16)) / (self.durability * 0.33 + self.speed * 0.33 + self.intelligence * 0.16 + self.strength * 0.16)
			damage *= ratio
		target.health = max(0, round(target.health - damage, 1)) # health will not drop below zero
		

class Game():
	def __init__(self, player_one, player_two):
		self.p1 = player_one
		self.p2 = player_two
		self.turn = -1
		self.offense = None
		self.defense = None
		#self.p1.print_stats()
		#self.p2.print_stats()
		self.ratio = 1
		if (not self.p1.health):
			self.ratio = float('inf')
		else:
			self.ratio =  self.p2.health / self.p1.health
	def simulate(self, depth):
		if (self.turn == -1):
			self.turn = 0 if ((self.p1.speed * 0.5 + self.p1.intelligence * 0.5) >= (self.p2.speed * 0.5 + self.p2.intelligence * 0.5)) else 1
		# 0 if p1 goes first, 1 if p2 goes first
		#print("Depth remaining:\t" + str(depth) + "\t Healths left:\t"  + str(self.p1.health) + " " + str(self.p2.health) + "\tType:\t" + str(self.p1.blocking) + " " + str(self.p2.blocking))
		if (self.p1.health <= 0):
			#print("Btw, looks like p2 won!")
			return
		if (self.p2.health <= 0):
			#print("Btw, looks like p1 won!")
			return
		self.ratio =  self.p2.health / self.p1.health
		#print (self.p2.health / self.p1.health)
		if (depth > 0):
			self.offense = deepcopy(self)
			self.defense = deepcopy(self)
			self.offense.turn = 0 if self.turn else 1
			self.defense.turn = 0 if self.turn else 1
			if (self.turn == 0):
				self.defense.p1.strike(self.offense.p2, 1)
				self.offense.p1.strike(self.defense.p2, 0)
			else:
				self.defense.p2.strike(self.defense.p1, 1)
				self.offense.p2.strike(self.offense.p1, 0)
			self.offense.simulate(depth - 1)
			self.defense.simulate(depth - 1)
			
class ExpectimaxHeuristic():
	def __init__(self, game):
		self.p1 = game.p1
		self.p2 = game.p2
		self.game_state = game
	def emulate(self):
		self.game_state.simulate(2)
		#print(self.game_state.offense.offense.ratio)
		#print(self.game_state.offense.defense.ratio)
		#print(self.game_state.defense.offense.ratio)
		#print(self.game_state.defense.defense.ratio)
		#print("\n" + str(self.game_state.offense.ratio))
		#print(self.game_state.defense.ratio)
		off_ratio = 0
		def_ratio = 0
		
		if (self.game_state.offense == None or self.game_state.defense == None):
			return
		if (self.game_state.offense.offense == None or self.game_state.offense.defense == None or self.game_state.defense.offense == None or self.game_state.defense.defense == None):
			off_ratio = self.game_state.offense.ratio
			def_ratio = self.game_state.defense.ratio
		else:
			off_ratio = (0.5 * self.game_state.offense.offense.ratio) + (0.5 * self.game_state.offense.defense.ratio)
			def_ratio = (0.5 * self.game_state.defense.offense.ratio) + (0.5 * self.game_state.defense.defense.ratio)
		#print("\n" + str(off_ratio))
		#print(def_ratio)
		active_player = self.p2 if self.game_state.turn else self.p1
		new_game = None
		attack_type = None
		if (active_player.player_type):
			if (off_ratio >= def_ratio):
				new_game = self.game_state.offense
				attack_type = " offensively."
			else:
				new_game = self.game_state.defense
				attack_type = " defensively."
		else:
			if (off_ratio < def_ratio):
				new_game = self.game_state.offense
				attack_type = " offensively."
			else:
				new_game = self.game_state.defense
				attack_type = " defensively."
		if self.game_state.turn:
			print(str(self.p2.super_name) + " strikes " + str(self.p1.super_name) + attack_type)
		else:
			print(str(self.p1.super_name) + " strikes " + str(self.p2.super_name) + attack_type)
		print("Health: " + str(new_game.p1.health) + " " + str(new_game.p2.health) + "\n")
		if (not new_game.p2.health):
			print("Player 1 asserts dominance.")
			return
		elif (not new_game.p1.health):
			print("Player 2 asserts dominance.")
			return
		new_expheur = ExpectimaxHeuristic(new_game)
		new_expheur.emulate()
		return

main()
